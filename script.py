# coding: UTF-8

#モジュール読み込み
import openpyxl

#デスクトップ上の既存Exelファイルの取得
wb = openpyxl.load_workbook('/Users/kicks-t73/Desktop/test.xlsx')

#ワークシートを新規作成("名前")
#wb.create_sheet("new")

#シート読み込み
#ws = wb.active

#「Sheet2」作成
wb.create_sheet('Sheet2')

#それぞれのシートを読み込む
ws = wb['Sheet1']
ws2 = wb['Sheet2']

#Sheet1のB3セルの値をSheet2のB3セルにコピーする
#ws2["B2"] = ws["B2"].value

#「Sheet3」作成
#wb.create_sheet('Sheet3')

#それぞれのシートを読み込む
#ws2 = wb['Sheet2']
#ws3 = wb['Sheet3']

#Sheet2のB3セルの値をSheet3のB3セルにコピーする
#ws3["B2"] = ws2["B2"].value

#Sheet3のA1にNo.、B1に長さ、C1に幅を入力
#ws3.cell(row = 1, column = 1).value = "No."
#ws3.cell(row = 1, column = 2).value = "長さ(mm)"
#ws3.cell(row = 1, column = 3).value = "幅(mm)"


#b1 = ws['B1'].value
#b2 = ws['B2'].value
#b3 = ws['B3'].value
#「row = 行数」、「column = 列数」、「value = 値」
#ws.cell(row = 1, column = 3).value = b1*10 
#ws.cell(row = 2, column = 3).value = b2*10
#ws.cell(row = 3, column = 3).value = b3#*10


#別シートへ複数セルをコピー&ペースト
def copy_paste1():
    for i in range(2, ws.max_row + 1):
        number = ws.cell(row = i, column = 2).value
        ws2.cell(row = i, column = 2, value = number)
        i += 1
copy_paste1()


#B列を取得
#col = ws['B']

#B列の全値を出力
#for col_data in col:
    #print(col_data.value)

#保存
wb.save('/Users/kicks-t73/Desktop/test.xlsx')
wb.close()

